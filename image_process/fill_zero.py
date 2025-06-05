from openpyxl import load_workbook

# Đường dẫn đến file Excel
file_path = "D:\Research\GAT_SBIR\dataset\ChairV2\ChairV2_labels.xlsx"

# Mở file Excel
workbook = load_workbook(filename=file_path)
sheet = workbook.active  # Hoặc workbook["Tên sheet"] nếu bạn biết rõ

# Duyệt qua phạm vi A1 đến T401
for row in sheet.iter_rows(min_row=1, max_row=401, min_col=1, max_col=20):
    for cell in row:
        if cell.value is None:
            cell.value = 0

# Lưu lại file đã chỉnh sửa
workbook.save("D:\Research\GAT_SBIR\dataset\ChairV2\ChairV2_labels_filled.xlsx")
